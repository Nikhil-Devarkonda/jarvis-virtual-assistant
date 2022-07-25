import tkinter as tk
import tkinter.font as tkFont
entry_list = []
window = tk.Tk()
def save():
    global entry_list,window
    # Execl
    import openpyxl as xl
    wb = xl.load_workbook('Data/Contacts.xlsx')
    sheet = wb.active
    last_row = sheet.max_row + 1
    for i,ent in enumerate(entry_list):
        print(chr(65+i)+str(last_row),i,ent)
        sheet[chr(65+i)+str(last_row)].value = ent.get()
    wb.save('Contacts.xlsx')
    print('Saved')
    window.destroy()

def get_contacts(name):
    import openpyxl as xl
    wb = xl.load_workbook('Data/Contacts.xlsx')
    sheet = wb.active
    max_cols = sheet.max_column
    max_row = sheet.max_row
    contact_list = []
    for i in range(1,sheet.max_row+1):
        if name.lower() in sheet['A'+str(i)].value.lower() :
            contact = []
            for j in range(1,sheet.max_column+1):
                contact.append(sheet[chr(64+j)+str(i)].value)
            contact_list.append(contact)
    print(contact_list)
    

    

window.geometry("520x200")
window.title('Add Contacts')
font = tkFont.Font(family="Lucida Grande", size=13)
##font  = tk.Font()
for i,name in enumerate(['Name ','Email  ','Mobile No  ','D.O.B(DD/MM/YYYY)  ','Address  ','Relation ']):
    frame  = tk.Frame(master=window)
    lbl_name = tk.Label(master=frame,text=name,width=20,font=font)
    lbl_name.pack(fill=tk.X,side=tk.LEFT)
    entry_list.append(tk.Entry(master=frame,width=35,font=font))
    entry_list[-1].pack(fill=tk.X,side=tk.RIGHT)
    frame.pack()
btn = tk.Button(master=window,text='Save',command=save,font=font)
btn.pack()
window.mainloop()



